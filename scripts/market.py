#!/usr/bin/env python3
"""
CLI למשיכת נתוני עסקאות שוק לפי אזור.

שימוש:
    python -m scripts.market "רמת גן"
    python -m scripts.market "פלורנטין תל אביב" --pages 5 --out data/
"""
from __future__ import annotations

import argparse
import csv
import datetime
import sys
from pathlib import Path

# מאפשר הרצה ישירה מתיקיית השורש
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from realestate.nadlan import get_market_data  # noqa: E402


def export_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def export_excel(rows: list[dict], summary: dict, area: str, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    su = wb.active
    su.title = "סיכום"
    su.sheet_view.rightToLeft = True
    su["A1"] = f"נתוני שוק — {area}"
    su["A1"].font = Font(bold=True, size=14, color="1F4E78")
    su["A2"] = f"נמשך בתאריך {datetime.date.today():%d/%m/%Y} ממאגר נדל\"ן הממשלתי"
    su["A2"].font = Font(italic=True, size=10, color="808080")
    labels = {
        "deals": "מספר עסקאות", "avg_price_per_sqm": "מחיר ממוצע למ\"ר",
        "median_price_per_sqm": "מחיר חציוני למ\"ר", "avg_price": "מחיר עסקה ממוצע",
        "median_price": "מחיר עסקה חציוני", "min_price": "עסקה זולה", "max_price": "עסקה יקרה",
    }
    r = 4
    for k, v in summary.items():
        su[f"A{r}"] = labels.get(k, k)
        su[f"A{r}"].font = Font(bold=True)
        su[f"B{r}"] = v
        su[f"B{r}"].number_format = "#,##0"
        su[f"B{r}"].fill = PatternFill("solid", fgColor="DDEBF7")
        r += 1
    su.column_dimensions["A"].width = 26
    su.column_dimensions["B"].width = 16

    dt = wb.create_sheet("עסקאות")
    dt.sheet_view.rightToLeft = True
    if rows:
        cols = list(rows[0].keys())
        for j, c in enumerate(cols, 1):
            cell = dt.cell(row=1, column=j, value=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for i, row in enumerate(rows, 2):
            for j, c in enumerate(cols, 1):
                dt.cell(row=i, column=j, value=row[c])
        for col in "ABCDEFGHI":
            dt.column_dimensions[col].width = 14
    wb.save(path)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="משיכת נתוני עסקאות נדל\"ן לפי אזור")
    ap.add_argument("area", help="שם עיר או שכונה, למשל \"רמת גן\"")
    ap.add_argument("--pages", type=int, default=3, help="מספר עמודי עסקאות למשיכה")
    ap.add_argument("--out", default=".", help="תיקיית פלט")
    args = ap.parse_args(argv)

    print(f'מחפש עסקאות עבור "{args.area}" ...')
    try:
        rows, summary = get_market_data(args.area, args.pages)
    except Exception as e:  # noqa: BLE001
        print(f"שגיאה: {e}", file=sys.stderr)
        return 1
    if not rows:
        print("לא נמצאו עסקאות.", file=sys.stderr)
        return 1

    print(f"נמצאו {summary['deals']} עסקאות.")
    for k, v in summary.items():
        print(f"  {k}: {v:,}" if isinstance(v, (int, float)) else f"  {k}: {v}")

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    stem = args.area.replace(" ", "_")
    export_csv(rows, out / f"deals_{stem}.csv")
    export_excel(rows, summary, args.area, out / f"market_{stem}.xlsx")
    print(f"נשמר: {out / f'market_{stem}.xlsx'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
